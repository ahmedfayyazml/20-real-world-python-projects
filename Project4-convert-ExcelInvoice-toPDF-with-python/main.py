import pandas as pd
from fpdf import FPDF
import  glob
from pathlib import Path

from openpyxl.styles.builtins import total

filepaths = glob.glob("invoices/*.xlsx")

for filepath in filepaths:
    #set up pdf orientation
    pdf = FPDF(orientation="P",unit="mm",format="A4")
    # added page
    pdf.add_page()

    # added filename
    filename = Path(filepath).stem
    invoice_num,date = filename.split("-")

    #Number
    pdf.set_font(style="B",size=16,family="Times")
    pdf.cell(w=50,h=8,txt=f"Invoice nr{invoice_num}",ln=1)

    #date
    pdf.set_font(style="B", size=16, family="Times")
    pdf.cell(w=50, h=8, txt=f"Date : {date}",ln=1)


    # Collected data from dataframe
    df = pd.read_excel(filepath,"Sheet 1")

    #added heading
    columns = list(df.columns)
    pdf.set_font(family="Times", size=10,style="B")
    pdf.set_text_color(80, 80, 80)
    pdf.cell(h=8, w=30, txt=str(columns[0].replace("_"," ").title()), border=1)
    pdf.cell(h=8, w=65, txt=str(columns[1].replace("_"," ").title()), border=1)
    pdf.cell(h=8, w=35, txt=str(columns[2].replace("_"," ").title()), border=1)
    pdf.cell(h=8, w=30, txt=str(columns[3].replace("_"," ").title()), border=1)
    pdf.cell(h=8, w=30, txt=str(columns[4].replace("_"," ").title()), border=1, ln=1)

    #adding columns
    for index,row in df.iterrows():
        pdf.set_font(family="Times",size=10)
        pdf.set_text_color(80,80,80)
        pdf.cell(h=8,w=30,txt=str(row["product_id"]),border=1)
        pdf.cell(h=8,w=65,txt=str(row["product_name"]),border=1)
        pdf.cell(h=8,w=35,txt=str(row["amount_purchased"]),border=1)
        pdf.cell(h=8,w=30,txt=str(row["price_per_unit"]),border=1)
        pdf.cell(h=8,w=30,txt=str(row["total_price"]),border=1,ln=1)
    # added total sum
    total_sum = df["total_price"].sum()
    pdf.cell(h=8, w=30, txt="", border=1)
    pdf.cell(h=8, w=65, txt="", border=1)
    pdf.cell(h=8, w=35, txt="", border=1)
    pdf.cell(h=8, w=30, txt="", border=1)
    pdf.cell(h=8, w=30, txt=str(total_sum), border=1, ln=1)

    #writting total amount
    pdf.set_font(style="B",size=12,family="Times")
    pdf.cell(h=8,w=30,txt="The total sum is "+str(total_sum))
    pdf.output("PDFs/"+filename+".pdf")
